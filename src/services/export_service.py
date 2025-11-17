"""
导出服务
支持导出为 CSV、Excel、JSON、Anki 等格式
"""
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional
from loguru import logger

from src.data.models import Entry
from src.data.repository import EntryRepository


class ExportService:
    """导出服务"""

    def __init__(self):
        self.entry_repo = EntryRepository()

    def export_to_csv(
        self,
        output_path: str,
        entries: Optional[List[Entry]] = None,
        include_all_fields: bool = False
    ) -> bool:
        """
        导出为CSV格式

        Args:
            output_path: 输出文件路径
            entries: 词条列表（None则导出全部）
            include_all_fields: 是否包含所有字段

        Returns:
            是否成功
        """
        try:
            if entries is None:
                entries = self.entry_repo.get_all(limit=10000)

            if not entries:
                logger.warning("没有词条可导出")
                return False

            # 确定字段
            if include_all_fields:
                fieldnames = [
                    'id', 'source_text', 'translation', 'source_lang', 'target_lang',
                    'entry_type', 'context', 'source_app', 'source_url',
                    'familiarity', 'proficiency', 'review_count', 'correct_count',
                    'last_review', 'next_review', 'ease_factor', 'interval',
                    'is_starred', 'tags', 'notes', 'translator_type',
                    'translation_time', 'created_at', 'updated_at'
                ]
            else:
                fieldnames = [
                    'source_text', 'translation', 'pronunciation',
                    'entry_type', 'tags', 'notes', 'proficiency',
                    'review_count', 'created_at'
                ]

            # 写入CSV
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for entry in entries:
                    row = {}
                    for field in fieldnames:
                        value = getattr(entry, field, '')
                        # 格式化特殊字段
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif value is None:
                            value = ''
                        row[field] = value

                    writer.writerow(row)

            logger.info(f"成功导出 {len(entries)} 条记录到 {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            return False

    def export_to_excel(
        self,
        output_path: str,
        entries: Optional[List[Entry]] = None,
        include_all_fields: bool = False
    ) -> bool:
        """
        导出为Excel格式

        Args:
            output_path: 输出文件路径
            entries: 词条列表（None则导出全部）
            include_all_fields: 是否包含所有字段

        Returns:
            是否成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            if entries is None:
                entries = self.entry_repo.get_all(limit=10000)

            if not entries:
                logger.warning("没有词条可导出")
                return False

            # 构建数据
            data = []
            for entry in entries:
                if include_all_fields:
                    row = {
                        'ID': entry.id,
                        '原文': entry.source_text,
                        '翻译': entry.translation,
                        '源语言': entry.source_lang,
                        '目标语言': entry.target_lang,
                        '类型': entry.entry_type,
                        '上下文': entry.context,
                        '来源应用': entry.source_app,
                        '来源URL': entry.source_url,
                        '熟悉度': entry.familiarity,
                        '熟练度': entry.proficiency,
                        '复习次数': entry.review_count,
                        '正确次数': entry.correct_count,
                        '上次复习': entry.last_review,
                        '下次复习': entry.next_review,
                        '难度系数': entry.ease_factor,
                        '间隔天数': entry.interval,
                        '收藏': '是' if entry.is_starred else '否',
                        '标签': entry.tags,
                        '笔记': entry.notes,
                        '翻译器': entry.translator_type,
                        '翻译耗时': entry.translation_time,
                        '创建时间': entry.created_at,
                        '更新时间': entry.updated_at
                    }
                else:
                    row = {
                        '原文': entry.source_text,
                        '翻译': entry.translation,
                        '类型': entry.entry_type,
                        '标签': entry.tags,
                        '笔记': entry.notes,
                        '熟练度': entry.proficiency,
                        '复习次数': entry.review_count,
                        '创建时间': entry.created_at.strftime('%Y-%m-%d %H:%M:%S') if entry.created_at else ''
                    }
                data.append(row)

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 导出到Excel
            df.to_excel(output_path, index=False, engine='openpyxl')

            # 美化Excel
            wb = load_workbook(output_path)
            ws = wb.active

            # 设置标题样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output_path)

            logger.info(f"成功导出 {len(entries)} 条记录到 {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False

    def export_to_json(
        self,
        output_path: str,
        entries: Optional[List[Entry]] = None,
        pretty: bool = True
    ) -> bool:
        """
        导出为JSON格式

        Args:
            output_path: 输出文件路径
            entries: 词条列表（None则导出全部）
            pretty: 是否格式化输出

        Returns:
            是否成功
        """
        try:
            if entries is None:
                entries = self.entry_repo.get_all(limit=10000)

            if not entries:
                logger.warning("没有词条可导出")
                return False

            # 构建数据
            data = []
            for entry in entries:
                item = {
                    'id': entry.id,
                    'source_text': entry.source_text,
                    'translation': entry.translation,
                    'source_lang': entry.source_lang,
                    'target_lang': entry.target_lang,
                    'entry_type': entry.entry_type,
                    'context': entry.context,
                    'source_app': entry.source_app,
                    'source_url': entry.source_url,
                    'familiarity': entry.familiarity,
                    'proficiency': entry.proficiency,
                    'review_count': entry.review_count,
                    'correct_count': entry.correct_count,
                    'last_review': entry.last_review.isoformat() if entry.last_review else None,
                    'next_review': entry.next_review.isoformat() if entry.next_review else None,
                    'ease_factor': entry.ease_factor,
                    'interval': entry.interval,
                    'is_starred': entry.is_starred,
                    'tags': entry.tags,
                    'notes': entry.notes,
                    'translator_type': entry.translator_type,
                    'translation_time': entry.translation_time,
                    'created_at': entry.created_at.isoformat() if entry.created_at else None,
                    'updated_at': entry.updated_at.isoformat() if entry.updated_at else None
                }
                data.append(item)

            # 写入JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                else:
                    json.dump(data, f, ensure_ascii=False)

            logger.info(f"成功导出 {len(entries)} 条记录到 {output_path}")
            return True

        except Exception as e:
            logger.error(f"导出JSON失败: {e}")
            return False

    def export_to_anki(
        self,
        output_path: str,
        entries: Optional[List[Entry]] = None
    ) -> bool:
        """
        导出为Anki格式（txt文件，tab分隔）

        Args:
            output_path: 输出文件路径
            entries: 词条列表（None则导出全部）

        Returns:
            是否成功
        """
        try:
            if entries is None:
                entries = self.entry_repo.get_all(limit=10000)

            if not entries:
                logger.warning("没有词条可导出")
                return False

            # 写入Anki格式
            with open(output_path, 'w', encoding='utf-8') as f:
                for entry in entries:
                    # Anki格式：正面	背面	标签
                    front = entry.source_text
                    back = entry.translation

                    # 添加笔记到背面
                    if entry.notes:
                        back += f"\n\n<div class='notes'>{entry.notes}</div>"

                    # 标签
                    tags = entry.tags if entry.tags else ""

                    # 写入（tab分隔）
                    f.write(f"{front}\t{back}\t{tags}\n")

            logger.info(f"成功导出 {len(entries)} 条记录到 {output_path} (Anki格式)")
            return True

        except Exception as e:
            logger.error(f"导出Anki格式失败: {e}")
            return False

    def export(
        self,
        output_path: str,
        format: str = "csv",
        entries: Optional[List[Entry]] = None,
        **kwargs
    ) -> bool:
        """
        统一导出接口

        Args:
            output_path: 输出文件路径
            format: 导出格式 (csv/excel/json/anki)
            entries: 词条列表
            **kwargs: 其他参数

        Returns:
            是否成功
        """
        format = format.lower()

        if format == "csv":
            return self.export_to_csv(output_path, entries, **kwargs)
        elif format in ["excel", "xlsx"]:
            return self.export_to_excel(output_path, entries, **kwargs)
        elif format == "json":
            return self.export_to_json(output_path, entries, **kwargs)
        elif format == "anki":
            return self.export_to_anki(output_path, entries)
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False
